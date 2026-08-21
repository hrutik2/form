from datetime import datetime, timedelta, timezone
from io import BytesIO
from secrets import token_urlsafe
from zoneinfo import ZoneInfo

from bson import ObjectId
from bson.errors import InvalidId
from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pymongo import ReturnDocument
from pymongo.errors import PyMongoError

from app.core.config import settings
from app.db.mongo import db
from app.utils.security import create_form_submission_token, decode_form_submission_token

IST = ZoneInfo("Asia/Kolkata")


def serialize(document: dict | None):
    if not document:
        return None
    document["_id"] = str(document["_id"])
    return document


def with_detail(detail: str, **payload):
    return {"detail": detail, **payload}


def build_public_form_link(form_id: str, token: str | None = None) -> str:
    base_url = f"{settings.public_form_base_url.rstrip('/')}/{form_id}"
    if token:
        return f"{base_url}?token={token}"
    return base_url


def normalize_emails(emails: list[str]) -> list[str]:
    unique_emails: list[str] = []
    seen = set()
    for email in emails:
        normalized = email.strip().lower()
        if normalized and normalized not in seen:
            seen.add(normalized)
            unique_emails.append(normalized)
    return unique_emails


def parse_object_id(value: str) -> ObjectId:
    try:
        return ObjectId(value)
    except InvalidId as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid form id",
        ) from error


def raise_database_error(action: str, error: PyMongoError):
    raise HTTPException(
        status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
        detail=f"Database unavailable while trying to {action}",
    ) from error


def format_ist_datetime(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(IST).strftime("%d %b %Y || %I:%M:%S %p").lower()


def ensure_utc_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def list_forms():
    try:
        cursor = db.forms.find().sort("updated_at", -1).limit(200)
        return [serialize(item) for item in cursor]
    except PyMongoError as error:
        raise_database_error("list forms", error)


def get_form(form_id: str):
    try:
        return serialize(db.forms.find_one({"_id": parse_object_id(form_id)}))
    except PyMongoError as error:
        raise_database_error("fetch a form", error)


def create_form(payload: dict, created_by: str | None = None):
    now = datetime.now(timezone.utc)
    payload["created_at"] = now
    payload["updated_at"] = now
    payload["created_by"] = created_by
    payload["expiry_enabled"] = payload.get("expiry_enabled", False)
    payload["token_reuse_enabled"] = payload.get("token_reuse_enabled", False)
    try:
        result = db.forms.insert_one(payload)
        payload["_id"] = result.inserted_id
        return with_detail("Form created successfully", form=serialize(payload))
    except PyMongoError as error:
        raise_database_error("create a form", error)


def update_form(form_id: str, payload: dict):
    payload["updated_at"] = datetime.now(timezone.utc)
    try:
        updated = db.forms.find_one_and_update(
            {"_id": parse_object_id(form_id)},
            {"$set": payload},
            return_document=ReturnDocument.AFTER,
        )
        if not updated:
            return None
        return with_detail("Form updated successfully", form=serialize(updated))
    except PyMongoError as error:
        raise_database_error("update a form", error)


def delete_form(form_id: str):
    try:
        form = db.forms.find_one({"_id": parse_object_id(form_id)}, {"status": 1})
        if not form:
            return None
        if form.get("status") == "published":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Published forms cannot be deleted",
            )

        result = db.forms.delete_one({"_id": form["_id"]})
        return with_detail("Form deleted successfully", deleted=result.deleted_count == 1)
    except HTTPException:
        raise
    except PyMongoError as error:
        raise_database_error("delete a form", error)


def publish_form(
    form_id: str,
    published_by: str | None = None,
    enable_expiry: bool = False,
    expiry_value: int | None = None,
    expiry_unit: str | None = None,
    recipient_emails: list[str] | None = None,
    token_reuse_enabled: bool = False,
):
    now = datetime.now(timezone.utc)
    target_id = parse_object_id(form_id)
    expires_at = None
    normalized_emails = normalize_emails(recipient_emails or [])

    if enable_expiry:
        if expiry_value is None or not expiry_unit:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Expiry time is required when expiry is enabled",
            )
        if expiry_value <= 0:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Expiry time must be greater than zero",
            )

        duration_map = {
            "minutes": {"minutes": expiry_value},
            "hours": {"hours": expiry_value},
            "days": {"days": expiry_value},
            "weeks": {"weeks": expiry_value},
        }
        expires_at = now + timedelta(**duration_map[expiry_unit])
        if not normalized_emails:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Recipient emails are required when expiry is enabled",
            )

    try:
        db.forms.update_many(
            {"status": "published", "_id": {"$ne": target_id}},
            {
                "$set": {
                    "status": "unpublished",
                    "updated_at": now,
                    "expires_at": None,
                    "expiry_enabled": False,
                }
            },
        )
        published = db.forms.find_one_and_update(
            {"_id": target_id},
            {
                "$set": {
                    "status": "published",
                    "published_at": now,
                    "updated_at": now,
                    "expires_at": expires_at,
                    "expiry_enabled": enable_expiry,
                    "token_reuse_enabled": token_reuse_enabled,
                    "created_by": published_by,
                },
                "$inc": {"version": 1},
            },
            return_document=ReturnDocument.AFTER,
        )
        if not published:
            return None

        db.form_recipients.delete_many({"form_id": str(published["_id"])})
        recipient_links: list[dict] = []
        if enable_expiry:
            recipients = []
            for email in normalized_emails:
                token = token_urlsafe(24)
                recipients.append(
                    {
                        "form_id": str(published["_id"]),
                        "email": email,
                        "token": token,
                        "token_expiry": expires_at,
                        "token_status": "ACTIVE",
                        "created_at": now,
                        "submitted_at": None,
                    }
                )
                recipient_links.append(
                    {
                        "email": email,
                        "token": token,
                        "link": build_public_form_link(str(published["_id"]), token),
                        "token_expiry": expires_at,
                        "token_status": "ACTIVE",
                    }
                )
            if recipients:
                db.form_recipients.insert_many(recipients)

        return with_detail(
            "Form published successfully",
            form=serialize(published),
            recipient_links=recipient_links,
            public_link=build_public_form_link(str(published["_id"])),
        )
    except PyMongoError as error:
        raise_database_error("publish a form", error)


def get_published_form():
    try:
        form = db.forms.find_one({"status": "published"}, sort=[("updated_at", -1)])
    except PyMongoError as error:
        raise_database_error("fetch the published form", error)

    serialized = serialize(form)
    if not serialized:
        return None

    serialized["expires_at"] = ensure_utc_datetime(serialized.get("expires_at"))
    if serialized.get("expiry_enabled"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Form link token is required",
        )
    return serialized


def validate_public_form_access(form_id: str, token: str | None = None):
    form = get_form(form_id)
    if not form:
        return None
    if form.get("status") != "published":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Form not found")

    if not form.get("expiry_enabled"):
        return {"form": form, "recipient": None}

    if not token:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Form link token is required",
        )

    try:
        recipient = db.form_recipients.find_one({"form_id": form_id, "token": token})
    except PyMongoError as error:
        raise_database_error("validate public form access", error)

    if not recipient:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid form link",
        )

    token_expiry = ensure_utc_datetime(recipient.get("token_expiry"))
    if recipient.get("token_status") != "ACTIVE":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid form link",
        )
    if token_expiry and token_expiry <= datetime.now(timezone.utc):
        raise HTTPException(
            status_code=status.HTTP_410_GONE,
            detail="Form link expired",
        )

    return {"form": form, "recipient": serialize(recipient)}


def get_submission_token(form_id: str):
    try:
        form = db.forms.find_one(
            {"_id": parse_object_id(form_id)},
            {"status": 1, "expires_at": 1},
        )
    except PyMongoError as error:
        raise_database_error("fetch a submission token", error)

    if not form:
        return None
    if form.get("status") != "published":
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Form is not published",
        )

    expires_at = ensure_utc_datetime(form.get("expires_at"))
    if expires_at and expires_at <= datetime.now(timezone.utc):
        raise HTTPException(status_code=status.HTTP_410_GONE, detail="Form expired")

    return with_detail(
        "Submission token generated successfully",
        submission_token=create_form_submission_token(form_id, expires_at),
    )


def create_submission(
    form_id: str,
    payload: dict,
    submission_token: str | None,
    access_token: str | None = None,
    form_version: int | None = None,
):
    try:
        if form_version is None:
            form = db.forms.find_one(
                {"_id": parse_object_id(form_id)},
                {"version": 1, "status": 1, "expires_at": 1, "expiry_enabled": 1, "token_reuse_enabled": 1},
            )
            if not form:
                return None
            if form.get("status") != "published":
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Form is not published",
                )
            expires_at = ensure_utc_datetime(form.get("expires_at"))
            if expires_at and expires_at <= datetime.now(timezone.utc):
                raise HTTPException(status_code=status.HTTP_410_GONE, detail="Form expired")
            form_version = form["version"]
        else:
            form = db.forms.find_one(
                {"_id": parse_object_id(form_id)},
                {"version": 1, "status": 1, "expires_at": 1, "expiry_enabled": 1, "token_reuse_enabled": 1},
            )
            if not form:
                return None

        recipient_id = None
        submitted_by = None

        if form.get("expiry_enabled"):
            if not access_token:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Form link token is required",
                )
            try:
                recipient = db.form_recipients.find_one({"form_id": form_id, "token": access_token})
            except PyMongoError as error:
                raise_database_error("validate a form recipient", error)

            if not recipient:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid form link")
            token_expiry = ensure_utc_datetime(recipient.get("token_expiry"))
            if recipient.get("token_status") != "ACTIVE":
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid form link")
            if token_expiry and token_expiry <= datetime.now(timezone.utc):
                raise HTTPException(status_code=status.HTTP_410_GONE, detail="Form link expired")

            if not submission_token:
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail="Submission token is required",
                )
            try:
                token_payload = decode_form_submission_token(submission_token)
            except ValueError as error:
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail=str(error),
                ) from error

            if token_payload["sub"] != form_id:
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail="Invalid submission token",
                )

            recipient_id = str(recipient["_id"])
            submitted_by = recipient.get("email")
        elif submission_token:
            try:
                token_payload = decode_form_submission_token(submission_token)
            except ValueError as error:
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail=str(error),
                ) from error
            if token_payload["sub"] != form_id:
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail="Invalid submission token",
                )

        now = datetime.now(timezone.utc)
        submission = {
            "form_id": form_id,
            "form_version": form_version,
            "recipient_id": recipient_id,
            "submitted_by": submitted_by,
            "data": payload,
            "submitted_at": now,
        }
        result = db.submissions.insert_one(submission)
        submission["_id"] = str(result.inserted_id)
        if form.get("expiry_enabled") and recipient_id and not form.get("token_reuse_enabled"):
            db.form_recipients.update_one(
                {"_id": parse_object_id(recipient_id)},
                {"$set": {"token_status": "USED", "submitted_at": now}},
            )
        return with_detail("Form submitted successfully", submission=submission)
    except HTTPException:
        raise
    except PyMongoError as error:
        raise_database_error("create a submission", error)


def list_submissions(form_id: str):
    try:
        items = db.submissions.find({"form_id": form_id}).sort("submitted_at", -1).limit(1000)
        return [serialize(item) for item in items]
    except PyMongoError as error:
        raise_database_error("list submissions", error)


def export_submissions_xlsx(form_id: str):
    try:
        form = db.forms.find_one({"_id": parse_object_id(form_id)}, {"sections": 1})
        submissions = list(db.submissions.find({"form_id": form_id}).sort("submitted_at", -1).limit(5000))
    except PyMongoError as error:
        raise_database_error("export submissions", error)

    if not form:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Form not found")

    fields = [
        field
        for section in form["sections"]
        for row in section["rows"]
        for field in row["fields"]
    ]
    columns = [(field.get("name") or field["id"], field["label"]) for field in fields]

    workbook = Workbook()
    sheet = workbook.active
    headers = ["Submitted At", *[label for _, label in columns]]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for item in submissions:
        row = [format_ist_datetime(item["submitted_at"])]
        row.extend(item.get("data", {}).get(key, "") for key, _ in columns)
        sheet.append(row)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column_cells in enumerate(sheet.columns, start=1):
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
